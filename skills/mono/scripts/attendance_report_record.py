#!/usr/bin/env python3
"""
考勤记录报表导出脚本 — 补卡/出差/外出/请假

属于考勤报表导出体系，和 attendance_report_detail.py / attendance_report_monthly.py 平级。
Agent 负责意图判断和人员获取，本脚本自包含：数据查询 → 解析 → Excel 生成。

数据链路:
  1. dws attendance approve list --users <ids> --types <type> --start --end
     → 获取审批单摘要（含 originId = processInstanceId）
  2. dws oa approval detail --instance-id <originId>
     → 获取审批单完整表单字段（extValue / detailList）
  3. 解析 formValueVOS 中的 DDHolidayField / extValue → 按天拆分行
  4. write_excel 输出

用法:
  python attendance_report_record.py --type leave --users <userId1,userId2> --start 2026-04-01 --end 2026-04-30
  python attendance_report_record.py --type trip --users <userId1,userId2> --start 2026-04-01 --end 2026-04-30
  python attendance_report_record.py --type out --users <userId1> --start 2026-05-01 --end 2026-05-31
  python attendance_report_record.py --type patch --users <userId1> --start 2026-05-01 --end 2026-05-31

支持类型: leave(请假), trip(出差), out(外出), patch(补卡)
"""

from __future__ import annotations

import argparse
import json
import sys
import os
from datetime import datetime
from typing import Any

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from attendance_report_common import (
    run_dws,
    write_excel,
    resolve_user_names,
    resolve_user_info,
    UserInfo,
    log,
    warn,
    error,
    DwsCallError,
    DATE_FMT,
)

# ─────────────────────────────────────────────────────────────────────────────
# 常量
# ─────────────────────────────────────────────────────────────────────────────

SUPPORTED_TYPES = ("leave", "trip", "out", "patch")

COLUMNS: dict[str, list[str]] = {
    "leave": ["姓名", "考勤组", "部门", "工号", "职位", "假期类型", "请假时间",
              "请假时长(小时)", "请假时长(天)", "关联审批单", "审批单状态"],
    "trip": ["姓名", "考勤组", "部门", "工号", "职位", "出差时间",
             "出差时长", "出差单位", "关联审批单", "审批单状态"],
    "out": ["姓名", "考勤组", "部门", "工号", "职位", "外出申请时间",
            "外出时长(小时)", "外出时长(天)", "关联审批单", "审批单状态"],
    "patch": ["姓名", "考勤组", "部门", "工号", "职位", "考勤日期", "考勤时间",
              "原打卡时间", "原考勤状态", "补卡时间", "补卡结果", "关联审批单", "审批单状态"],
}

SHEET_NAMES: dict[str, str] = {
    "leave": "请假记录",
    "trip": "出差记录",
    "out": "外出记录",
    "patch": "补卡记录",
}

STATUS_MAP: dict[str, dict[str, str]] = {
    "COMPLETED": {"agree": "审批通过", "refuse": "已拒绝"},
    "RUNNING": {"": "审批中"},
    "TERMINATED": {"": "已撤销"},
}

APPROVE_LIST_BATCH_SIZE = 50  # attendance approve list 单次最多用户数

# 审批详情页 URL 模板
# 内层：aflow 审批详情页
_AFLOW_URL_TEMPLATE = (
    "https://aflow.dingtalk.com/dingtalk/mobile/homepage.htm"
    "?corpid={corp_id}&dd_share=false&showmenu=true&back=native"
    "#/approval?procInstId={instance_id}"
)
# 外层：dingtalk schema 协议，在钉钉客户端侧边面板打开
_DINGTALK_SCHEMA_TEMPLATE = (
    "dingtalk://dingtalkclient/action/openapp"
    "?corpid={corp_id}&container_type=slide_panel&app_id=-4"
    "&&redirect_url={encoded_url}"
)


class HyperlinkCell:
    """标记单元格为超链接：Excel 中显示 label 文本，点击跳转到 url。"""

    __slots__ = ("label", "url")

    def __init__(self, label: str, url: str):
        self.label = label
        self.url = url

    def __str__(self) -> str:
        return self.label


def build_approve_url(corp_id: str, instance_id: str) -> str:
    """
    构建审批单跳转链接（dingtalk:// schema）。

    结构：外层 dingtalk schema 打开钉钉侧边面板，内部 redirect 到 aflow 审批详情页。
    """
    from urllib.parse import quote

    inner_url = _AFLOW_URL_TEMPLATE.format(corp_id=corp_id, instance_id=instance_id)
    encoded_url = quote(inner_url, safe="")
    return _DINGTALK_SCHEMA_TEMPLATE.format(corp_id=corp_id, encoded_url=encoded_url)


def build_approve_cell(corp_id: str, instance_id: str, title: str = "") -> HyperlinkCell | str:
    """
    构建"关联审批单"列的单元格值。

    如果有 corp_id 和 instance_id，返回 HyperlinkCell（Excel 中为可点击链接）。
    否则返回纯文本。
    """
    if not instance_id:
        return ""
    label = title or instance_id
    if not corp_id:
        return label
    url = build_approve_url(corp_id, instance_id)
    return HyperlinkCell(label=label, url=url)


# ─────────────────────────────────────────────────────────────────────────────
# 工具函数
# ─────────────────────────────────────────────────────────────────────────────

def normalize_am_pm(text: str) -> str:
    """将时间文本中的 AM/PM 替换为 上午/下午。"""
    return text.replace(" PM", " 下午").replace(" AM", " 上午")


def format_status(status: str, result: str) -> str:
    """将 status + processInstanceResult 转为中文状态。"""
    status_upper = (status or "").upper()
    result_lower = (result or "").lower()
    group = STATUS_MAP.get(status_upper, {})
    return group.get(result_lower, group.get("", f"{status}/{result}"))


def ms_to_datetime(ms: int | float | None) -> datetime | None:
    """毫秒时间戳转 datetime。"""
    if not ms:
        return None
    try:
        return datetime.fromtimestamp(int(ms) / 1000)
    except (OSError, ValueError, OverflowError):
        return None


def ms_to_time_str(ms: int | float | None) -> str:
    """毫秒时间戳转 HH:MM。"""
    dt = ms_to_datetime(ms)
    return dt.strftime("%H:%M") if dt else ""


def ms_to_date_str(ms: int | float | None) -> str:
    """毫秒时间戳转 YYYY-MM-DD。"""
    dt = ms_to_datetime(ms)
    return dt.strftime(DATE_FMT) if dt else ""


def format_day_type(detail: dict) -> str:
    """从 detailList 单条判断日历类型。"""
    day_type = detail.get("dayType", "")
    is_rest = detail.get("isRest", False)
    if day_type == "workDay" or (not is_rest and not day_type):
        return "工作日"
    if day_type == "restDay" or is_rest:
        return "休息日"
    if day_type == "holiday":
        return "节假日"
    return day_type or ("休息日" if is_rest else "工作日")


def format_class_time(detail: dict) -> str:
    """从 detailList 单条提取上下班时间。"""
    class_info = detail.get("classInfo", {})
    sections = class_info.get("sections", []) if class_info else []
    if not sections:
        return "未排班"
    section = sections[0]
    start_time = ms_to_time_str(section.get("startTime"))
    end_time = ms_to_time_str(section.get("endTime"))
    if start_time and end_time:
        return f"{start_time} ~ {end_time}"
    return "未排班"


# ─────────────────────────────────────────────────────────────────────────────
# 数据查询
# ─────────────────────────────────────────────────────────────────────────────

# 钉钉接口把"外出"和"出差"都归类到 trip（bizType=2），out 类型查不到数据。
# 脚本通过 tagName 区分：tagName="出差" → trip，tagName="外出" → out。
_API_TYPE_MAP: dict[str, str] = {
    "leave": "leave",
    "trip": "trip",
    "out": "trip",   # 外出也用 trip 查询，再按 tagName 过滤
    "patch": "patch",
}

_TAG_FILTER: dict[str, str | None] = {
    "leave": None,
    "trip": "出差",
    "out": "外出",
    "patch": None,
}


def fetch_approve_list(user_ids: list[str], record_type: str, start: str, end: str) -> list[dict]:
    """
    分批调用 dws attendance approve list 获取审批单摘要。

    返回列表中每条包含: userId, tagName, duration, durationUnit, beginTime, endTime, originId。
    对于 out 类型，实际用 trip 查询接口，再按 tagName="外出" 过滤；
    对于 trip 类型，按 tagName="出差" 过滤（排除外出记录）。
    """
    api_type = _API_TYPE_MAP.get(record_type, record_type)
    tag_filter = _TAG_FILTER.get(record_type)

    all_records: list[dict] = []
    for i in range(0, len(user_ids), APPROVE_LIST_BATCH_SIZE):
        batch = user_ids[i:i + APPROVE_LIST_BATCH_SIZE]
        users_str = ",".join(batch)
        try:
            result = run_dws([
                "attendance", "approve", "list",
                "--users", users_str,
                "--types", api_type,
                "--start", start,
                "--end", end,
            ])
            records: list[dict] = []
            if isinstance(result, list):
                records = result
            elif isinstance(result, dict):
                records = result.get("approveList", result.get("list", []))
                if not isinstance(records, list):
                    records = []
            # 按 tagName 过滤
            if tag_filter:
                records = [r for r in records if r.get("tagName") == tag_filter]
            all_records.extend(records)
        except DwsCallError as e:
            warn(f"查询审批列表失败(batch {i // APPROVE_LIST_BATCH_SIZE + 1}): {e}")
    return all_records


def fetch_detail(instance_id: str) -> dict | None:
    """调用 dws oa approval detail 获取审批单完整详情。"""
    try:
        result = run_dws([
            "oa", "approval", "detail",
            "--instance-id", instance_id,
        ])
        return result if isinstance(result, dict) else None
    except DwsCallError as e:
        warn(f"获取审批详情失败({instance_id[:20]}...): {e}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
# 解析器
# ─────────────────────────────────────────────────────────────────────────────

def find_holiday_field(form_values: list[dict]) -> dict | None:
    """从 formValueVOS 中查找 DDHolidayField 组件。"""
    for fv in form_values:
        if fv.get("componentType") == "DDHolidayField":
            return fv
    return None


def parse_ext_value(field_data: dict) -> dict:
    """解析字段的 extValue JSON 字符串。"""
    ext_str = field_data.get("extValue") or ""
    if not ext_str:
        return {}
    try:
        return json.loads(ext_str)
    except (json.JSONDecodeError, TypeError):
        return {}


def parse_leave_detail(detail: dict, name_map: dict[str, str], *,
                       user_info_map: dict[str, "UserInfo"] | None = None,
                       group_map: dict[str, str] | None = None,
                       corp_id: str = "",
                       ) -> list[list[str]]:
    """解析请假审批单。"""
    form_values = detail.get("formValueVOS", [])
    user_id = detail.get("originatorUserid", "")
    dept_name = detail.get("originatorDeptName", "")
    instance_id = detail.get("processInstanceId", "")
    status = format_status(detail.get("status", ""), detail.get("processInstanceResult", ""))

    # 用户基础信息
    info = (user_info_map or {}).get(user_id)
    user_name = info.name if info else name_map.get(user_id, user_id)
    dept = info.dept_name if info and info.dept_name else dept_name
    job_number = info.job_number if info else ""
    title = info.title if info else ""
    group_name = (group_map or {}).get(user_id, "")
    approve_cell = build_approve_cell(corp_id, instance_id, f"{user_name}提交的请假审批单")

    holiday_field = find_holiday_field(form_values)
    if not holiday_field:
        return [[user_name, group_name, dept, job_number, title,
                 "", "", "", "", approve_cell, status]]

    # value: ["开始时间","结束时间",天数,"单位","假期类型","请假类型"]
    value_str = holiday_field.get("value", "")
    leave_type = ""
    leave_time = ""
    try:
        value_arr = json.loads(value_str)
        if isinstance(value_arr, list) and len(value_arr) >= 2:
            leave_time = normalize_am_pm(f"{value_arr[0]} ~ {value_arr[1]}")
            if len(value_arr) > 4:
                leave_type = str(value_arr[4])
    except (json.JSONDecodeError, TypeError):
        leave_time = normalize_am_pm(value_str)

    ext = parse_ext_value(holiday_field)
    duration_day = str(ext.get("durationInDay", ""))
    duration_hour = str(ext.get("durationInHour", ""))

    return [[user_name, group_name, dept, job_number, title,
             leave_type, leave_time, duration_hour, duration_day,
             approve_cell, status]]


def _extract_time_duration_from_fields(form_values: list[dict]) -> tuple[str, str, str, str]:
    """
    从独立表单字段中提取时间范围和时长。

    适用于外出/出差表单的非 DDHolidayField 结构：
      - startTime (DDDateField) + finishTime (DDDateField) → 时间范围
      - duration (NumberField) → extValue 中含 durationInDay / durationInHour

    Returns: (time_range, duration_hour, duration_day, ext_from_duration)
    """
    start_time = ""
    end_time = ""
    duration_hour = ""
    duration_day = ""

    for fv in form_values:
        biz_alias = (fv.get("bizAlias") or "").lower()
        name = (fv.get("name") or "").lower()
        value = fv.get("value") or ""

        # 开始时间
        if biz_alias in ("starttime", "start_time") or "开始时间" in name:
            if value and not start_time:
                start_time = value
        # 结束时间
        if biz_alias in ("finishtime", "finish_time", "endtime", "end_time") or "结束时间" in name:
            if value and not end_time:
                end_time = value
        # 时长字段 — extValue 中有 durationInDay / durationInHour
        if biz_alias == "duration" or "时长" in name:
            ext = parse_ext_value(fv)
            if ext:
                duration_day = str(ext.get("durationInDay", ""))
                duration_hour = str(ext.get("durationInHour", ""))

    time_range = ""
    if start_time and end_time:
        time_range = f"{start_time} ~ {end_time}"
    elif start_time:
        time_range = start_time

    return time_range, duration_hour, duration_day


def parse_out_detail(detail: dict, name_map: dict[str, str], *,
                     user_info_map: dict[str, "UserInfo"] | None = None,
                     group_map: dict[str, str] | None = None,
                     corp_id: str = "",
                     ) -> list[list[str]]:
    """解析外出审批单。兼容 DDHolidayField 和独立字段两种表单结构。"""
    form_values = detail.get("formValueVOS", [])
    user_id = detail.get("originatorUserid", "")
    dept_name = detail.get("originatorDeptName", "")
    instance_id = detail.get("processInstanceId", "")
    status = format_status(detail.get("status", ""), detail.get("processInstanceResult", ""))

    # 用户基础信息
    info = (user_info_map or {}).get(user_id)
    user_name = info.name if info else name_map.get(user_id, user_id)
    dept = info.dept_name if info and info.dept_name else dept_name
    job_number = info.job_number if info else ""
    title = info.title if info else ""
    group_name = (group_map or {}).get(user_id, "")
    approve_cell = build_approve_cell(corp_id, instance_id, f"{user_name}提交的外出审批单")

    # 优先尝试 DDHolidayField
    holiday_field = find_holiday_field(form_values)
    if holiday_field:
        value_str = holiday_field.get("value", "")
        time_range = ""
        try:
            value_arr = json.loads(value_str)
            if isinstance(value_arr, list) and len(value_arr) >= 2:
                time_range = normalize_am_pm(f"{value_arr[0]} ~ {value_arr[1]}")
        except (json.JSONDecodeError, TypeError):
            time_range = normalize_am_pm(value_str)
        ext = parse_ext_value(holiday_field)
        duration_day = str(ext.get("durationInDay", ""))
        duration_hour = str(ext.get("durationInHour", ""))
    else:
        # 回退: 从独立字段提取
        time_range, duration_hour, duration_day = _extract_time_duration_from_fields(form_values)

    return [[user_name, group_name, dept, job_number, title,
             time_range, duration_hour, duration_day,
             approve_cell, status]]


def parse_trip_from_approve_record(
    record: dict,
    name_map: dict[str, str],
    *,
    user_info_map: dict[str, "UserInfo"] | None = None,
    group_map: dict[str, str] | None = None,
    corp_id: str = "",
) -> list[str]:
    """
    直接从 attendance approve list 的记录中解析出差行。

    不依赖 oa approval detail（该接口对出差单存在 saNode 类型冲突 bug），
    仅使用 approve list 返回的 beginTime/endTime/duration/durationUnit/originId。
    """
    user_id = record.get("userId", "")
    info = (user_info_map or {}).get(user_id)
    user_name = info.name if info else name_map.get(user_id, user_id)
    dept = info.dept_name if info else ""
    job_number = info.job_number if info else ""
    title = info.title if info else ""
    group_name = (group_map or {}).get(user_id, "")

    begin_ms = record.get("beginTime")
    end_ms = record.get("endTime")
    begin_str = ms_to_date_str(begin_ms) if begin_ms else ""
    end_str = ms_to_date_str(end_ms) if end_ms else ""
    time_range = f"{begin_str} ~ {end_str}" if begin_str and end_str else begin_str or end_str

    duration = record.get("duration", "")
    duration_unit = record.get("durationUnit", "DAY")
    unit_str = "天" if duration_unit == "DAY" else "小时"

    instance_id = record.get("originId", "")
    effective_corp_id = corp_id or record.get("corpId", "")
    approve_cell = build_approve_cell(effective_corp_id, instance_id, f"{user_name}提交的出差审批单")

    # approve list 没有审批状态，有 gmtFinished 说明已完结，视为审批通过
    status = "审批通过" if record.get("gmtFinished") else "审批中"

    return [user_name, group_name, dept, job_number, title,
            time_range, str(duration), unit_str, approve_cell, status]


def fetch_check_results(user_ids: list[str], start: str, end: str) -> dict[str, list[dict]]:
    """
    批量查询打卡结果，返回 {userId: [records...]} 映射。

    每条 record 含: workDate, timeResult, planCheckTime, userCheckTime 等。
    """
    result_map: dict[str, list[dict]] = {}
    batch_size = 50
    for i in range(0, len(user_ids), batch_size):
        batch = user_ids[i:i + batch_size]
        try:
            result = run_dws([
                "attendance", "check", "result",
                "--users", ",".join(batch),
                "--start", start,
                "--end", end,
            ])
            records = []
            if isinstance(result, list):
                records = result
            elif isinstance(result, dict):
                records = result.get("result", result.get("list", []))
                if not isinstance(records, list):
                    records = []
            for rec in records:
                uid = rec.get("userId", "")
                if uid:
                    result_map.setdefault(uid, []).append(rec)
        except DwsCallError as e:
            warn(f"查询打卡结果失败(batch {i // batch_size + 1}): {e}")
    return result_map


def fetch_user_group_map(user_ids: list[str]) -> dict[str, str]:
    """
    查询考勤组列表并建立 userId → 考勤组名称映射。

    流程：先 group search 拿到所有考勤组 ID+名称，
    再对有成员的考勤组调用 filtered-get --member 获取成员列表。
    """
    group_map: dict[str, str] = {}
    user_id_set = set(user_ids)

    try:
        result = run_dws(["attendance", "group", "search"])
        items: list[dict] = []
        if isinstance(result, list):
            items = result
        elif isinstance(result, dict):
            # 适配 {items: [...]} 或 {result: {items: [...]}}
            inner = result.get("items", result.get("result", result))
            if isinstance(inner, dict):
                items = inner.get("items", [])
            elif isinstance(inner, list):
                items = inner

        for g in items:
            group_name = g.get("name", g.get("groupName", ""))
            group_id = g.get("id", g.get("groupId", ""))
            member_count = g.get("memberCount", 0)

            if not group_id or not group_name or not member_count:
                continue

            # 调用 filtered-get 获取成员 userId 列表
            try:
                detail = run_dws([
                    "attendance", "group", "filtered-get",
                    "--group-id", str(group_id), "--member",
                ])
                member_users: list[str] = []
                if isinstance(detail, dict):
                    member_users = detail.get("memberUsers", [])
                    if not isinstance(member_users, list):
                        member_users = []
                for uid in member_users:
                    uid_str = str(uid)
                    if uid_str in user_id_set:
                        group_map[uid_str] = group_name
            except DwsCallError:
                pass

    except DwsCallError as e:
        warn(f"查询考勤组失败: {e}")
    return group_map


CHECK_TIME_RESULT_MAP = {
    "Normal": "正常",
    "Late": "迟到",
    "Early": "早退",
    "Absenteeism": "旷工",
    "NotSigned": "未打卡",
    "SeriousLate": "严重迟到",
}


def parse_patch_detail(detail: dict, name_map: dict[str, str], *,
                       user_info_map: dict[str, "UserInfo"] | None = None,
                       group_map: dict[str, str] | None = None,
                       check_result_map: dict[str, list[dict]] | None = None,
                       corp_id: str = "",
                       ) -> list[list[str]]:
    """解析补卡审批单，输出完整列。"""
    form_values = detail.get("formValueVOS", [])
    user_id = detail.get("originatorUserid", "")
    dept_name = detail.get("originatorDeptName", "")
    instance_id = detail.get("processInstanceId", "")
    status = format_status(detail.get("status", ""), detail.get("processInstanceResult", ""))

    # 用户基础信息
    info = (user_info_map or {}).get(user_id)
    user_name = info.name if info else name_map.get(user_id, user_id)
    dept = info.dept_name if info and info.dept_name else dept_name
    job_number = info.job_number if info else ""
    title = info.title if info else ""
    group_name = (group_map or {}).get(user_id, "")
    approve_cell = build_approve_cell(corp_id, instance_id, f"{user_name}提交的补卡审批单")

    # 从表单解析补卡时间和原因
    patch_time = ""
    patch_reason = ""
    work_date = ""
    check_time_str = ""
    ext_data: dict = {}

    for fv in form_values:
        comp_type = fv.get("componentType", "") or ""
        biz_alias = (fv.get("bizAlias") or "").lower()
        name = fv.get("name") or ""
        value = fv.get("value") or ""

        if comp_type == "DDDateField" or "checktime" in biz_alias or "补卡时间" in name:
            if value and not patch_time:
                patch_time = value
            # 解析 extValue 获取考勤日期等
            ext = parse_ext_value(fv)
            if ext and not ext_data:
                ext_data = ext
        if "reason" in biz_alias or "原因" in name or "事由" in name or "理由" in name:
            if value and not patch_reason:
                patch_reason = value

    # 从 extValue 提取考勤日期、考勤时间、原考勤状态
    plan_tip = ""
    plan_text = ""
    if ext_data:
        work_date_ms = ext_data.get("workDate")
        if work_date_ms:
            work_date = ms_to_date_str(work_date_ms)
        plan_tip = ext_data.get("planTip", "")
        plan_text = ext_data.get("planText", "")

    # 从 planText / planTip 提取考勤时间（目标格式：YYYY-MM-DD HH:MM）
    # planText 格式: "2026-04-25,星期六,195固定班次,上班时间09:00"
    # planTip 格式: "周六上班(04.25 09:00) 缺卡" / "周一上班(04.27 09:00) 缺卡"
    import re
    plan_time_hhmm = ""

    # 优先从 planTip 提取（更可靠，含具体日期和时间）
    # planTip 格式: "周三下班(03.05 01:00) 缺卡"
    plan_date_from_tip = ""  # MM.DD → 用于跨日场景
    if plan_tip:
        # 匹配 "(MM.DD HH:MM)" 格式
        tip_match = re.search(r"\((\d{2})\.(\d{2})\s+(\d{2}:\d{2})\)", plan_tip)
        if tip_match:
            plan_date_from_tip = f"{tip_match.group(1)}-{tip_match.group(2)}"  # "03-05"
            plan_time_hhmm = tip_match.group(3)

    # 回退：从 planText 中提取
    if not plan_time_hhmm and plan_text:
        # 匹配 "上班时间HH:MM" 或 "下班时间HH:MM" 或 "时间HH:MM"
        time_match = re.search(r"时间(\d{2}:\d{2})", plan_text)
        if time_match:
            plan_time_hhmm = time_match.group(1)

    # 最后回退：任意 HH:MM 格式
    if not plan_time_hhmm:
        for source in (plan_tip, plan_text):
            if source:
                fallback_match = re.search(r"(\d{2}:\d{2})", source)
                if fallback_match:
                    plan_time_hhmm = fallback_match.group(1)
                    break

    # 拼接考勤时间：优先使用 planTip 中解析的完整日期（处理跨日班次）
    if plan_date_from_tip and plan_time_hhmm and work_date:
        # 用 work_date 的年份 + planTip 中的 MM-DD + HH:MM
        year = work_date[:4]
        check_time_str = f"{year}-{plan_date_from_tip} {plan_time_hhmm}"
    elif work_date and plan_time_hhmm:
        check_time_str = f"{work_date} {plan_time_hhmm}"
    elif plan_time_hhmm:
        check_time_str = plan_time_hhmm
    elif plan_text:
        check_time_str = plan_text
    elif plan_tip:
        check_time_str = plan_tip

    # 如果 work_date 为空，从 patch_time 中提取日期
    if not work_date and patch_time:
        work_date = patch_time[:10] if len(patch_time) >= 10 else ""

    # 从 planTip / planText 提取原考勤状态
    # planTip 格式: "周六上班(04.25 09:00) 缺卡" / "Thursday ( 04.23 ) Adjust"
    # planText 格式: "2026-04-25,星期六,195固定班次,上班时间09:00" 或 "周一上班(04.27 09:00) 缺卡"
    original_check_time = ""
    original_status = ""

    tip_status_map = {
        "缺卡": "缺卡", "未打卡": "未打卡",
        "迟到": "迟到", "早退": "早退",
        "旷工": "旷工", "正常": "正常",
        "NotSigned": "未打卡", "Adjust": "已调整",
    }
    # 优先从 planTip 提取，回退到 planText
    for source in (plan_tip, plan_text):
        if source:
            for keyword, label in tip_status_map.items():
                if keyword in source:
                    original_status = label
                    break
            if original_status:
                break

    # 回退: 尝试从 check result 接口获取（如果有数据）
    if check_result_map and user_id in check_result_map:
        for rec in check_result_map[user_id]:
            rec_date = rec.get("workDate", "")
            if isinstance(rec_date, (int, float)):
                rec_date = ms_to_date_str(rec_date)
            if rec_date == work_date:
                user_check_ms = rec.get("userCheckTime")
                if user_check_ms:
                    dt = ms_to_datetime(user_check_ms)
                    original_check_time = dt.strftime("%Y-%m-%d %H:%M") if dt else ""
                time_result = rec.get("timeResult", "")
                if time_result:
                    original_status = CHECK_TIME_RESULT_MAP.get(time_result, time_result)
                break

    # 补卡结果：审批通过 → 补卡成功
    patch_result = ""
    if status == "审批通过":
        patch_result = "补卡成功"
    elif status == "已拒绝":
        patch_result = "补卡失败"
    elif status == "审批中":
        patch_result = "待审批"
    elif status == "已撤销":
        patch_result = "已撤销"

    return [[user_name, group_name, dept, job_number, title, work_date, check_time_str,
             original_check_time, original_status, patch_time, patch_result,
             approve_cell, status]]


PARSERS = {
    "leave": parse_leave_detail,
    "out": parse_out_detail,
    "patch": parse_patch_detail,
}

# 需要额外用户信息（考勤组/工号/职位）的类型
_TYPES_NEED_USER_INFO = {"leave", "out", "patch", "trip"}


# ─────────────────────────────────────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="考勤记录报表导出（补卡/出差/外出/请假）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--type", required=True, choices=SUPPORTED_TYPES,
                        help="记录类型: leave(请假)/trip(出差)/out(外出)/patch(补卡)")
    parser.add_argument("--users", required=True,
                        help="用户 ID 列表，逗号分隔（由 Agent 从人员获取阶段提供）")
    parser.add_argument("--start", required=True,
                        help="开始日期 YYYY-MM-DD")
    parser.add_argument("--end", required=True,
                        help="结束日期 YYYY-MM-DD")
    parser.add_argument("--out", default="",
                        help="输出文件路径（不传则自动生成）")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    record_type: str = args.type
    user_ids = [u.strip() for u in args.users.split(",") if u.strip()]
    start_date: str = args.start
    end_date: str = args.end

    if not user_ids:
        error("--users 不能为空")
        sys.exit(1)

    try:
        datetime.strptime(start_date, DATE_FMT)
        datetime.strptime(end_date, DATE_FMT)
    except ValueError:
        error("日期格式错误，请使用 YYYY-MM-DD")
        sys.exit(1)

    sheet_name = SHEET_NAMES[record_type]
    log(f"开始导出{sheet_name}：{len(user_ids)} 人，{start_date} ~ {end_date}")

    # ── Step 1: 获取审批单列表 ──
    log("步骤 1/4：查询审批单列表...")
    approve_records = fetch_approve_list(user_ids, record_type, start_date, end_date)
    log(f"  获取到 {len(approve_records)} 条审批记录")

    if not approve_records:
        log("未查询到任何记录")
        print(f"{sheet_name}：0 条记录，无需生成文件")
        sys.exit(0)

    # 从 approve list 记录中提取 corpId（用于构建审批单跳转链接）
    corp_id = ""
    for r in approve_records:
        if r.get("corpId"):
            corp_id = r["corpId"]
            break

    # ── Step 2: 去重提取 instanceId ──
    instance_ids = list(dict.fromkeys(
        r.get("originId", "") for r in approve_records if r.get("originId")
    ))
    log(f"步骤 2/4：共 {len(instance_ids)} 个审批实例")

    # ── Step 3: 解析用户信息 ──
    log("步骤 3/4：解析用户信息...")
    name_map = resolve_user_names(user_ids)

    user_info_map: dict[str, UserInfo] | None = None
    group_map: dict[str, str] | None = None
    check_result_map: dict[str, list[dict]] | None = None

    if record_type in _TYPES_NEED_USER_INFO:
        log("  获取用户完整信息（工号/职位）...")
        user_info_map = resolve_user_info(user_ids)
        log("  查询考勤组映射...")
        group_map = fetch_user_group_map(user_ids)

    if record_type == "patch":
        log("  查询原打卡结果...")
        check_result_map = fetch_check_results(user_ids, start_date, end_date)

    all_rows: list[list[str]] = []

    if record_type == "trip":
        # 出差记录直接从 approve list 数据生成，不调用 oa approval detail
        # （oa approval detail 对出差单存在 saNode result 字段类型冲突 bug）
        log("步骤 4/4：从审批列表解析出差记录...")
        for record in approve_records:
            row = parse_trip_from_approve_record(
                record, name_map,
                user_info_map=user_info_map,
                group_map=group_map,
                corp_id=corp_id,
            )
            all_rows.append(row)
    else:
        log("步骤 4/4：查询审批详情并解析...")
        for idx, instance_id in enumerate(instance_ids):
            if (idx + 1) % 10 == 0:
                log(f"  进度: {idx + 1}/{len(instance_ids)}")

            detail = fetch_detail(instance_id)
            if not detail:
                continue

            # 补充新发现的用户
            originator = detail.get("originatorUserid", "")
            if originator and originator not in name_map:
                extra = resolve_user_names([originator])
                name_map.update(extra)
            if originator and user_info_map and originator not in user_info_map:
                extra_info = resolve_user_info([originator])
                user_info_map.update(extra_info)

            if record_type == "patch":
                rows = parse_patch_detail(
                    detail, name_map,
                    user_info_map=user_info_map,
                    group_map=group_map,
                    check_result_map=check_result_map,
                    corp_id=corp_id,
                )
            elif record_type == "leave":
                rows = parse_leave_detail(
                    detail, name_map,
                    user_info_map=user_info_map,
                    group_map=group_map,
                    corp_id=corp_id,
                )
            elif record_type == "out":
                rows = parse_out_detail(
                    detail, name_map,
                    user_info_map=user_info_map,
                    group_map=group_map,
                    corp_id=corp_id,
                )
            else:
                rows = PARSERS[record_type](detail, name_map)
            all_rows.extend(rows)

    log(f"  解析完成，共 {len(all_rows)} 行")

    if not all_rows:
        log("无有效数据行")
        print(f"{sheet_name}：解析后 0 行有效数据，无需生成文件")
        sys.exit(0)

    # ── 写入 Excel ──
    out_path = args.out or f"attendance_report_record_{record_type}_{start_date}_{end_date}.xlsx"
    headers = COLUMNS[record_type]
    title = f"{sheet_name}  统计日期：{start_date} 至 {end_date}"
    subtitle = f"报表生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # 将 HyperlinkCell 转为纯文本供 write_excel 写入，之后再补超链接
    plain_rows = []
    hyperlink_cells: list[tuple[int, int, str]] = []  # (row_offset, col_idx, url)
    for row_offset, row in enumerate(all_rows):
        plain_row = []
        for col_idx, cell in enumerate(row):
            if isinstance(cell, HyperlinkCell):
                plain_row.append(cell.label)
                hyperlink_cells.append((row_offset, col_idx, cell.url))
            else:
                plain_row.append(cell)
        plain_rows.append(plain_row)

    write_excel(
        out_path,
        headers,
        plain_rows,
        sheet_name=sheet_name,
        title=title,
        subtitle=subtitle,
    )

    # 补充超链接
    if hyperlink_cells:
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(out_path)
        ws = wb.active
        # 计算标题行偏移：title + subtitle + header
        title_row_count = (1 if title else 0) + (1 if subtitle else 0)
        first_data_row = title_row_count + 2  # +1 for header, +1 for 1-indexed

        link_font = Font(color="0563C1", underline="single")
        for row_offset, col_idx, url in hyperlink_cells:
            cell = ws.cell(row=first_data_row + row_offset, column=col_idx + 1)
            cell.hyperlink = url
            cell.font = link_font
        wb.save(out_path)

    abs_path = os.path.abspath(out_path)
    log(f"✅ 导出完成: {abs_path}")
    print(f"{sheet_name}导出完成：{abs_path}（{len(all_rows)} 行，{len(instance_ids)} 个审批单）")


if __name__ == "__main__":
    main()
